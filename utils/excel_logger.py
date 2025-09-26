"""Helpers for appending call activity into Excel logs."""
from __future__ import annotations

import os
from datetime import datetime
from typing import Iterable, Optional

import csv
import re

try:
    from openpyxl import Workbook, load_workbook  # type: ignore
except ImportError:  # pragma: no cover
    Workbook = None  # type: ignore
    load_workbook = None  # type: ignore


_HEADER = [
    "timestamp",
    "user_id",
    "agent_id",
    "lead_name",
    "lead_phone",
    "call_status",
    "lead_status",
    "failure_reason",
    "key_phrases",
    "call_id",
]


_INVALID_CHARS = re.compile(r"[\x00-\x08\x0b-\x0c\x0e-\x1f]")


def _clean(value: Optional[str]) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r", " ").replace("\n", " ")
    return _INVALID_CHARS.sub("", text)


def _ensure_directory(path: str) -> None:
    os.makedirs(path, exist_ok=True)


def _initialise_workbook(path: str) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "calls"
    worksheet.append(_HEADER)
    workbook.save(path)


def log_call_summary(
    directory: str,
    user_id: str,
    agent_id: str,
    lead_name: str,
    lead_phone: str,
    call_status: str,
    lead_status: str,
    failure_reason: Optional[str],
    key_phrases: Iterable[str],
    call_id: str,
    timestamp: Optional[datetime] = None,
) -> str:
    """Append a call summary row and return the workbook path."""
    timestamp = timestamp or datetime.utcnow()
    _ensure_directory(directory)
    filename = os.path.join(directory, f"call_logs_{user_id}.xlsx")

    if not os.path.exists(filename):
        _initialise_workbook(filename)

    workbook = load_workbook(filename)
    worksheet = workbook.active
    worksheet.append(
        [
            timestamp.isoformat(),
            user_id,
            agent_id,
            lead_name,
            lead_phone,
            call_status,
            lead_status,
            _clean(failure_reason),
            _clean(", ".join(key_phrases)),
            call_id,
        ]
    )
    workbook.save(filename)
    return os.path.abspath(filename)
